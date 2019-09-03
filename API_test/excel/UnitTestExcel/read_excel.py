# -*- coding:utf-8 -*-
# @Time : 2019-08-24
# @Author：lqc
# @Email:572948875@qq.com
# File : read_excel.py

# eval()函数  把数据类型转换成原本的数据类型
s = "True"
eval(s)
a = '{"age":18}'
eval(s)

# 一次性读取所有数据 对内存要求高点
# 用的时候读取所有数据 磁盘读写要求高点
# 磁盘->内存->CPU

from openpyxl import load_workbook


class ReadExcel:
    def __init__(self, maxRow):
        self.maxRow = maxRow

    def readExcel(self, row, column):
        # 1、打开excel
        wb = load_workbook("unit_test_excel.xlsx")

        # 2、定位表单
        sheet = wb["python"]

        # 3、定位单元格 行列值去定位
        res = sheet.cell(row, column).value
        # 错误演示-：res = sheet.cell(row, column) <Cell 'Python' A1> 只定位到单元格，并未取值
        print("拿到的结果是:", res)


class DoExcel:
    def __init__(self, file_name, sheet_name):
        self.file_name = file_name
        self.sheet_name = sheet_name
        self.sheet_obj = load_workbook(self.file_name)[self.sheet_name]

    def get_header(self):
        # 获取第一行的标题行
        wb = load_workbook(self.file_name)
        sheet = wb[self.sheet_name]

        header = []  # 存储标题行
        for j in range(1, sheet.max_column + 1):
            header.append(sheet.cell(1, j).value)
        return header

    def get_data(self):
        # 根据嵌套循环读取数据
        wb = load_workbook(self.file_name)
        sheet = wb[self.sheet_name]
        test_data = []
        header = self.get_header()
        for i in range(2, sheet.max_row + 2):
            sub_data = {}
            for j in range(1, sheet.max_column + 1):
                sub_data[header[j-1]] = sheet.cell(i, j).value
            # sub_data = {'method': sheet.cell(i, 1).value,
            #             'url': sheet.cell(i, 1).value,
            #             'data': sheet.cell(i, 1).value,
            #             'expected': sheet.cell(i, 1).value}

            test_data.append(sub_data)
            return test_data
