# -*- coding:utf-8 -*-
# @Time : 2019-08-27
# @Author：lqc
# @Email:572948875@qq.com
# File : do_excel.py

from openpyxl import load_workbook
from PythonStudy.interface_auto_practice_50.tools.read_config import ReadConfig
from PythonStudy.interface_auto_practice_50.tools import project_path
from PythonStudy.interface_auto_practice_50.tools.get_data import GetData


# 相对路径 绝对路径
# wb = load_workbook(r"/Users/LiQingChun/Desktop/PythonStudy/PythonStudy/interface_auto_practice_50/test_data/interface_auto_practice.xlsx")
# wb1 = load_workbook("../test_data/interface_auto_practice.xlsx")


class DoExcel:
    def __init__(self, file_name):
        self.file_name = file_name

    # def get_headers(self):
    #     wb = load_workbook(self.file_name)
    #     sheet = wb[self.sheet_name]
    #     headers = []
    #     for i in range(1, sheet.max_column + 1):
    #         headers.append(sheet.cell(1, i).value)
    #     return headers

    # @classmethod
    def do_excel(self):
        wb = load_workbook(self.file_name)
        mode = eval(ReadConfig.get_config(project_path.case_config_path, 'MODE', 'mode'))

        tel = getattr(GetData, 'NoRegTel')  # 假设是从Excel里面拿到的手机号

        # 利用python查询数据库的方式，去拿到最大的手机号-这里可以加 也可以放到get_data

        # headers = cls.get_headers()
        result = []

        for key in mode:  # 遍历这个存在配置文件里面的字典
            sheet = wb[key]  # 表单名
            if mode[key] == 'all':
                for i in range(2, sheet.max_row + 1):
                    row_data = {}
                    row_data["case_id"] = sheet.cell(i, 1).value
                    row_data["url"] = sheet.cell(i, 2).value
                    # data["data"] = sheet.cell(i, 5).value
                    if sheet.cell(i,3).vaule.find('${tel_1}') != -1:
                        row_data["data"] = sheet.cell(i, 3).value.replace('${tel_1}', str(tel))
                    elif sheet.cell(i,3).vaule.find('${tel}') != -1:
                        row_data["data"] = sheet.cell(i, 3).value.replace('${tel}', str(tel+1))
                    else:
                        row_data["data"] = sheet.cell(i, 3).value
                    row_data["description"] = sheet.cell(i, 4).value
                    row_data["method"] = sheet.cell(i, 5).value
                    row_data["expected"] = sheet.cell(i, 6).value
                    row_data["sheet_name"] = key
                    # for j in range(1, sheet.max_column + 1):
                    #     data[headers[j - 1]] = sheet.cell(i, j).value
                    result.append(row_data)
                    self.update_tel(tel + 2, self.file_name, 'init')
            else:
                for case_id in mode[key]:
                    row_data = {}
                    row_data["case_id"] = sheet.cell(case_id + 1, 1).value
                    row_data["url"] = sheet.cell(case_id + 1, 2).value
                    # data["data"] = sheet.cell(case_id + 1, 5).value
                    if sheet.cell(case_id + 1, 3).vaule.find('${tel_1}') != -1:
                        row_data["data"] = sheet.cell(case_id + 1, 3).value.replace('${tel_1}', str(tel))
                    elif sheet.cell(case_id + 1, 3).vaule.find('${tel}') != -1:
                        row_data["data"] = sheet.cell(case_id + 1, 3).value.replace('${tel}', str(tel+1))
                    else:
                        row_data["data"] = sheet.cell(case_id + 1, 3).value
                    row_data["description"] = sheet.cell(case_id + 1, 4).value
                    row_data["method"] = sheet.cell(case_id + 1, 5).value
                    row_data["expected"] = sheet.cell(case_id + 1, 6).value
                    row_data["sheet_name"] = key
                    # for j in range(1, sheet.max_column + 1):
                    #     data[headers[j - 1]] = sheet.cell(i, j).value
                    result.append(row_data)
                    self.update_tel(tel + 2, self.file_name, 'init')

        return result

    @staticmethod
    def write_back(file_name, sheet_name, i, result, TestResult):
        wb = load_workbook(file_name)
        sheet = wb[sheet_name]
        sheet.cell(i, 7).value = result
        sheet.cell(i, 8).value = TestResult
        wb.save(file_name)  # 保存结果

    def update_tel(self, tel, filename, sheetname):
        # wb['init'].cell(2, 1).value = tel
        # wb.save()
        wb = load_workbook(filename)
        sheet = wb[sheetname]
        sheet.cell(2, 1).value = tel
        wb.save(filename)


if __name__ == "__main__":
    # /Users/szdl/Desktop/PythonStudy/PythonStudy/interface_auto_practice_50/test_data/interface_auto_practice.xlsx
    # /Users/szdl/Desktop/PythonStudy/PythonStudy/interface_auto_practice_50/run.py
    print(DoExcel(project_path.test_case_path, "python").do_excel())
