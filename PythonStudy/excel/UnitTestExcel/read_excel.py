# -*- coding:utf-8 -*-
# @Time : 2019-08-24
# @Author：lqc
# @Email:572948875@qq.com
# File : read_excel.py

# eval()函数  把数据类型转换成原本的数据类型
s = "True"
eval(s)
a = '{"age":18}'
eval(s)

from openpyxl import load_workbook


class ReadExcel:
    def __init__(self, maxRow):
        self.maxRow = maxRow

    def readExcel(self, row, column):
        # 1、打开excel
        wb = load_workbook("unit_test_excel.xlsx")

        # 2、定位表单
        sheet = wb["python"]

        # 3、定位单元格 行列值去定位
        res = sheet.cell(row, column).value
        # 错误演示-：res = sheet.cell(row, column) <Cell 'Python' A1> 只定位到单元格，并未取值
        print("拿到的结果是:", res)


class DoExcel:
    def __init__(self, file_name, sheet_name):
        self.file_name = file_name
        self.sheet_name = sheet_name
        self.sheet_obj = load_workbook(self.file_name)[self.sheet_name]

    def get_data(self):
        wb = load_workbook(self.file_name)
        sheet = wb[self.sheet_name]
        test_data = []
        for i in range(1, sheet.max_row + 1):
            sub_data = {'method': sheet.cell(i, 1).value,
                        'url': sheet.cell(i, 1).value,
                        'data': sheet.cell(i, 1).value,
                        'expected': sheet.cell(i, 1).value}

            test_data.append(sub_data)
            return test_data
