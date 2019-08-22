# -*- coding:utf-8 -*-
# @Time : 2019-08-22
# @Author：lqc
# @Email:572948875@qq.com
# File : excel.py

# python去操作excel
# 1、.xlsx openpyxl 只支持这种格式
# 2、
from openpyxl import load_workbook

# 1、打开excel
wb = load_workbook("python_excel.xlsx")  # Open the given filename and return the workbook

# 2、定位表单
sheet = wb["python"]

# 3、定位单元格 行列值去定位
res = sheet.cell(1, 1).value
print("拿到的结果是：", res)

print("最大行：{}".format(sheet.max_row))
print("最大列：{}".format(sheet.max_column))

# 数字还是数字，其他的都是字符串
print("url:{}, 类型是{}".format(sheet.cell(1, 1).value, type(sheet.cell(1, 1).value)))
print("url:{}, 类型是{}".format(sheet.cell(1, 1).value, type(sheet.cell(1, 1).value)))
print("url:{}, 类型是{}".format(sheet.cell(1, 1).value, type(sheet.cell(1, 1).value)))
print("url:{}, 类型是{}".format(sheet.cell(1, 1).value, type(sheet.cell(1, 1).value)))

# 请完成数据参数化+结合Excel
# 1、把数据存到excel
# 2、利用openpyxl写一个专门读取excel里面测试数据的类
# 3、结合单元测试方法，通过舒适化函数传参的方法，完成单元测试
# 4、操作excel的类，test_suite、test_http类
