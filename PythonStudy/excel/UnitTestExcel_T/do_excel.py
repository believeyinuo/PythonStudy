# -*- coding:utf-8 -*-
# @Time : 2019-08-25
# @Author：lqc
# @Email:572948875@qq.com
# File : do_excel.py

from openpyxl import load_workbook

class DoExcel:
    def __init__(self, file_name, sheet_name):
        self.file_name = file_name
        self.sheet_name = sheet_name

    def get_data(self):
        wb = load_workbook(self.file_name)
        sheet = wb[self.sheet_name]

        # 错误演示一：只定位到单元格，并未取值
        # res = sheet.cell(1, 1)
        # print(res)  # <Cell 'python' A1>

        test_data = []
        for i in range(1, sheet.max_row+1):
            sub_data = {'method': sheet.cell(i, 1).value,
                        'url': sheet.cell(1, 2).value,
                        'data': sheet.cell(1, 3).value,
                        'expected': sheet.cell(1, 4).value}
            test_data.append(sub_data)
        return test_data

