# -*- coding: utf-8 -*-
# @Time : 2019-08-26 12:23
# @Author : szdl
# @Email : believeyinuo@163.com
# @File : read_excel.py
# @Project : API_AUTO_test

from openpyxl import load_workbook
from PythonStudy.config_46.home_work.read_config import ReadConfig


class ReadExcel:
    def __init__(self, file_name, sheet_name):
        self.file_name = file_name
        self.sheet_name = sheet_name

    def get_header(self):
        wb = load_workbook(self.file_name)
        sheet = wb[self.sheet_name]
        header = []
        for j in range(1, sheet.max_column + 1):
            header.append(sheet.cell(1, j).value)
        return header

    def get_data(self):
        wb = load_workbook(self.file_name)
        sheet = wb[self.sheet_name]

        mode = ReadConfig.read_config("case.config", "MODE", 'mode')

        test_data = []

        header = self.get_header()

        for i in range(2, sheet.max_row + 1):
            sub_data = {}
            for j in range(1, sheet.max_column + 1):
                sub_data[header[j - 1]] = sheet.cell(i, j).value
            test_data.append(sub_data)

        if mode == 'all':
            final_data = test_data
        else:
            final_data = []
            for item in test_data:
                if item["case_id"] in mode:
                    final_data.append(item)

        return final_data
